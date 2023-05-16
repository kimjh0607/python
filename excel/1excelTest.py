import openpyxl as xl

wbook = xl.Workbook()
sheet = wbook.active
sheet["A1"] = "일찍 일어나는 새가 벌레를 잡는다"
sheet.cell(row=2, column=1, value="하늘은 스스로 돕는 자를 돕는다")
third_cell = sheet.cell(row=3, column=1)
third_cell.value = "낙숫물이 바위를 뚫는다"

for i in range(10):
    sheet.cell(row=(i+1), column=3, value=i)

sheet2 = wbook.create_sheet("구구단")
for y in range(1,10):
    for x in range(1,10):
        wcell = sheet2.cell(row=y, column=x)
        wcell.value = x * y

sheet3 = wbook.create_sheet("99x99")
for y in range(1,100):
    for x in range(1,100):
        wcell = sheet3.cell(y, x)
        wcell.value = x * y

wbook.save("./output/excelWrite.xlsx")

rbook = xl.load_workbook("./output/excelWrite.xlsx")

sheet = rbook.worksheets[0]
cell = sheet["A1"]
print(cell.value)
sheet2 = rbook.worksheets[1]
print(sheet2.cell(row=2,column=8).value)

for y in range(2, 5):
    r = []
    for x in range(2, 5):
        v = sheet2.cell(row=y, column=x).value
        r.append(v)
    print(r)

for row in sheet2["B3":"D5"]:
    r = []
    for cell in row:
        r.append(cell.value)
    print(r)
#list comprehension(for문 내에서 리스트를 생성할 때)
for row in sheet2["B4":"D6"]:
    print([c.value for c in row])
    
it = sheet2.iter_rows(
        min_row=2, min_col=2,
        max_row=4, max_col=4)

for row in it:
    print([c.value for c in row]) 
    # r = []
    # for cell in row:
    #     r.append(cell.value)
    # print(r)

cell = sheet2["C2"]
(row, col) = (cell.row, cell.column)
print("C2=({},{})".format(row, col))

cell = sheet2.cell(row=2, column=3)
cdt = cell.coordinate
print("(2,3)={}".format(cdt))