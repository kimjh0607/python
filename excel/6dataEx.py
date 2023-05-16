import openpyxl as xl

book = xl.load_workbook("D:\pythonwork\input\monthly_sales.xlsx")
sheet = book.active

rows = sheet["A3":"F9"]
for row in rows:
    values = [cell.value for cell in row]
    print(values)
    # values = [] 
    # for cell in row:
    #     values.append(cell.value)
    # print(values)

book2 = xl.load_workbook("input/monthly_sales.xlsx", data_only=True)
sheet2 = book2.active

rows2 = sheet2["A3":"F9"]
for row2 in rows2:
    values2 = [cell2.value for cell2 in row2]
    print(values2)
    
sheet3 =  xl.load_workbook("input/monthly_sales.xlsx",data_only=True).active
print((sheet3.max_row, sheet3.max_column))

rows = sheet3["A3":"F999"]
for row in rows:
    values = [cell.value for cell in row]
    if values[0] is None: break
    print(values)
    
for row in sheet3.iter_rows(min_row=3):
    values = [cell.value for cell in row]
    if values[0] is None: break
    print(values)
